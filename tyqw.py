"""
通义千问VL模型自动化测试（目前仅支持.xlsx文件）
@author: YTSakura
"""
import argparse
import os
import time

import openpyxl

from tyqwVL import aliyun

date = time.strftime("%Y%m%d-%H%M%S")


def read_excel(excel_file):
    """
    读取Excel文件内容
    :param excel_file: 文件路径
    :return:
    """
    book = ''
    if excel_file.endswith('xls'):
        print('暂不支持xls格式')
        exit(0)
    elif excel_file.endswith('xlsx'):
        book = openpyxl.load_workbook(excel_file)
    return book


def getsheetname(excel_file):
    """
    获取页签名称
    :param excel_file:
    :return:
    """
    sheetname = ''
    book = read_excel(excel_file)
    if excel_file.endswith('xls'):
        print('暂不支持xls格式')
        exit(0)
    elif excel_file.endswith('xlsx'):
        sheetname = book.sheetnames
    return sheetname, book


def get_value(sheetname, book):
    sheet = book[sheetname[0]]
    nrows = sheet.max_row
    ncols = sheet.max_column
    if sheet.cell(1, 1).value != 'question' or sheet.cell(1, 2).value != 'pic' or sheet.cell(1, 3).value != 'time':
        print('表格标题错误，请确认。')

    # 获取内容
    for i in range(2, nrows + 1):
        question.append(sheet.cell(i, 1).value)
        pic.append(sheet.cell(i, 2).value)
        times.append(sheet.cell(i, 3).value)


if __name__ == '__main__':

    parser = argparse.ArgumentParser(description='tyqw')
    parser.add_argument('--file', type=str, help='excel file path')
    parser.add_argument('--module', type=str, help='通义千问的模型')
    args = parser.parse_args()
    file = args.file
    module = args.module

    question, pic, times, result = [], [], [], []
    sheetname, book = getsheetname(file)

    if module == 'max':
        module = 'qwen-vl-max'
    elif module == 'plus':
        module = 'qwen-vl-plus'
    else:
        print('模型输入错误，请重新输入')
        exit(0)

    # 检测报告文件夹是否存在
    isExists = os.path.exists('./history')
    if not isExists:
        os.makedirs('./history')

    get_value(sheetname, book)

    for i in range(len(question)):
        if question[i] is None or pic[i] is None or times[i] is None:
            continue
        result.append(f"\n'{question[i]}'的{times[i]}次测试结果：")
        print(f"\n'{question[i]}'的{times[i]}次测试结果：")
        for _ in range(times[i]):
            test_result = aliyun(pic[i], question[i], module)
            time.sleep(1)
            result.append(test_result)

    report_file = open(fr'history\{date}-{module}-report.txt', 'w', encoding='utf-8')
    for item in result:
        report_file.write(str(item) + '\n')
